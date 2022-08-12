#Import Libraries
import requests
import openpyxl

#Function to generate url endpoint for api
def GetEndpoints():
    apis = []
    baseEndpoint = "https://fantasy.premierleague.com/api/leagues-h2h-matches/league/511906/"
    for i in range(1, 8):
        if i == 1 :
          print(baseEndpoint)
          apis.append(baseEndpoint)
        else:
            apiEndpoint = baseEndpoint + "?page=" + str(i)
            print(apiEndpoint)
            apis.append(apiEndpoint)
    return apis

#Function to get responses from API
def GetAPIResponse(APIEndpointURL):
    response = requests.get(APIEndpointURL).json()
    return response
    
#Convert API response to JSON
def ConvertToJson(data):
    response = data.json()

#Get all response jsons and write to excel
def GetAndWriteDataToExcel():
    apis = GetEndpoints()
    for api in apis:
        response = GetAPIResponse(api)
        writeJsonToExcel(sheet1,response)

#Function to remove repetition
def writeJsonToExcel(sheet1,json_history):
  global rownum
  for each in json_history["results"]:
    ide = each['id']
    Gameweek = each['event']
    Home = each['entry_1_name']
    Away = each['entry_2_name']
    HomePoints = each['entry_1_points']
    AwayPoints = each['entry_2_points']
    HomeTotal = each['entry_1_total']
    AwayTotal = each['entry_2_total']

    Rowvalue = sheet1.cell(row=2, column=3).value
    if not Rowvalue :
        rownum = 2

    sheet1.cell(row=rownum, column=3).value = ide
    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Home
    sheet1.cell(row=rownum, column=6).value = HomePoints
    sheet1.cell(row=rownum, column=7).value = HomeTotal
    rownum +=1

    sheet1.cell(row=rownum, column=4).value = Gameweek
    sheet1.cell(row=rownum, column=5).value = Away
    sheet1.cell(row=rownum, column=6).value = AwayPoints
    sheet1.cell(row=rownum, column=7).value = AwayTotal

    rownum +=1


#Running all of the functionality from here
#Create excel workbook,create sheets
global g_w,wb,sheet0,sheet1  
wb = openpyxl.Workbook()
sheet0 = wb.create_sheet(index=0, title='Read_Me')
sheet1 = wb.create_sheet(index=1, title='2019_2020')

#Create Read me sheet
sheet0['B2'].value = 'Welcome to FPL data'

header1 = ['id', 'GW', 'TeamName', 'GWPoints','H2HPts',]
headerrow = 1
for key in range(5):
    sheet1.cell(row=headerrow, column=key + 3).value = str(header1[key])
# Import gameweek history and insert data in sheet

#Get all apis, get data from each and write to excel
GetAndWriteDataToExcel()
#Save Workbook
wb.save("FPL.xlsx")
